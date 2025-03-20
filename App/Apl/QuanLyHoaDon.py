import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime
from tkcalendar import DateEntry
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
FORMAT_CURRENCY_VND = '#,##0 ₫'
def ket_noi_csdl():
    return sqlite3.connect("../Data/BTL_QLPT.db")

def get_next_bill_id():
    conn = ket_noi_csdl()
    cursor = conn.cursor()
    cursor.execute("SELECT MaHoaDon FROM HoaDon ORDER BY MaHoaDon DESC LIMIT 1")
    last_id = cursor.fetchone()
    conn.close()

    if last_id:
        num = int(last_id[0]) + 1
    else:
        num = 1
    return f"{num:03d}"  # Định dạng 3 chữ số: 001, 002, ...


# Hàm lấy giá dịch vụ điện/nước từ bảng DichVu
def get_service_prices():
    conn = ket_noi_csdl()
    cursor = conn.cursor()
    cursor.execute("SELECT TenDV, GiaDV FROM DichVu WHERE TenDV IN ('Điện', 'Nước')")
    prices = cursor.fetchall()
    conn.close()

    price_dict = {row[0]: row[1] for row in prices}
    return {'dien': price_dict.get('Điện', 0), 'nuoc': price_dict.get('Nước', 0)}


def quan_ly_hoa_don(parent_frame):
    for widget in parent_frame.winfo_children():
        widget.destroy()

    # Frame cho các chức năng
    frame_chuc_nang = tk.Frame(parent_frame, pady=10)
    frame_chuc_nang.pack(fill=tk.X)

    # Frame cho việc tìm kiếm
    frame_tim_kiem = tk.LabelFrame(frame_chuc_nang, text="Tìm kiếm", padx=5, pady=5)
    frame_tim_kiem.pack(side=tk.LEFT, padx=10)

    tk.Label(frame_tim_kiem, text="Mã hóa đơn:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
    entry_ma_hoa_don_search = tk.Entry(frame_tim_kiem, width=15)
    entry_ma_hoa_don_search.grid(row=0, column=1, padx=5, pady=5)

    tk.Label(frame_tim_kiem, text="Từ ngày:").grid(row=0, column=2, padx=5, pady=5, sticky=tk.W)
    entry_tu_ngay = DateEntry(frame_tim_kiem, width=12, background='darkblue', foreground='white',
                              date_pattern='dd/MM/yyyy')
    entry_tu_ngay.grid(row=0, column=3, padx=5, pady=5)

    tk.Label(frame_tim_kiem, text="Đến ngày:").grid(row=0, column=4, padx=5, pady=5, sticky=tk.W)
    entry_den_ngay = DateEntry(frame_tim_kiem, width=12, background='darkblue', foreground='white',
                               date_pattern='dd/MM/yyyy')
    entry_den_ngay.grid(row=0, column=5, padx=5, pady=5)

    btn_tim_kiem = tk.Button(frame_tim_kiem, text="Tìm kiếm",
                             command=lambda: search_bills(entry_ma_hoa_don_search.get(), entry_tu_ngay.get_date(),
                                                          entry_den_ngay.get_date(), tree_hoa_don))
    btn_tim_kiem.grid(row=0, column=6, padx=5, pady=5)

    # Frame cho các nút chức năng
    frame_buttons = tk.Frame(frame_chuc_nang)
    frame_buttons.pack(side=tk.RIGHT, padx=10)

    btn_them = tk.Button(frame_buttons, text="Thêm Hóa Đơn", command=lambda: show_bill_form(None, tree_hoa_don))
    btn_them.pack(side=tk.LEFT, padx=5)

    btn_sua = tk.Button(frame_buttons, text="Sửa", command=lambda: edit_selected_bill(tree_hoa_don))
    btn_sua.pack(side=tk.LEFT, padx=5)

    btn_xoa = tk.Button(frame_buttons, text="Xóa", command=lambda: delete_selected_bill(tree_hoa_don))
    btn_xoa.pack(side=tk.LEFT, padx=5)

    btn_xuat_excel = tk.Button(frame_buttons, text="Xuất Excel", command=lambda: export_to_excel(tree_hoa_don))
    btn_xuat_excel.pack(side=tk.LEFT, padx=5)

    btn_refresh = tk.Button(frame_buttons, text="Làm mới", command=lambda: refresh_data(tree_hoa_don))
    btn_refresh.pack(side=tk.LEFT, padx=5)

    # Frame cho bảng hiển thị hóa đơn
    frame_hoa_don = tk.Frame(parent_frame)
    frame_hoa_don.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    # Tạo Treeview để hiển thị hóa đơn
    columns = ('MaHoaDon', 'MaHD', 'MaPhong', 'CCCD', 'HoTen', 'SoTien', 'NgayLap', 'GiaPhong')
    tree_hoa_don = ttk.Treeview(frame_hoa_don, columns=columns, show='headings')

    # Định nghĩa tiêu đề các cột
    tree_hoa_don.heading('MaHoaDon', text='Mã Hóa Đơn')
    tree_hoa_don.heading('MaHD', text='Mã Hợp Đồng')
    tree_hoa_don.heading('MaPhong', text='Mã Phòng')
    tree_hoa_don.heading('NgayLap', text='Ngày Lập')
    tree_hoa_don.heading('SoTien', text='Số Tiền')
    tree_hoa_don.heading('HoTen', text='Họ Tên Người Thuê')
    tree_hoa_don.heading('CCCD', text='CCCD')
    tree_hoa_don.heading('GiaPhong', text='Giá Phòng')

    # Định nghĩa độ rộng các cột
    tree_hoa_don.column('MaHoaDon', width=100, anchor=tk.CENTER)
    tree_hoa_don.column('MaHD', width=100, anchor=tk.CENTER)
    tree_hoa_don.column('MaPhong', width=100, anchor=tk.CENTER)
    tree_hoa_don.column('NgayLap', width=150, anchor=tk.CENTER)
    tree_hoa_don.column('SoTien', width=150, anchor=tk.CENTER)
    tree_hoa_don.column('HoTen', width=200, anchor=tk.W)
    tree_hoa_don.column('CCCD', width=120, anchor=tk.CENTER)
    tree_hoa_don.column('GiaPhong', width=150, anchor=tk.CENTER)

    # Thêm thanh cuộn
    scrollbar = ttk.Scrollbar(frame_hoa_don, orient=tk.VERTICAL, command=tree_hoa_don.yview)
    tree_hoa_don.configure(yscroll=scrollbar.set)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    tree_hoa_don.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # Đổ dữ liệu vào Treeview
    refresh_data(tree_hoa_don)

    # Panel hiển thị chi tiết hóa đơn khi click vào một hóa đơn
    frame_chi_tiet = tk.LabelFrame(parent_frame, text="Chi tiết hóa đơn", padx=10, pady=10)
    frame_chi_tiet.pack(fill=tk.X, padx=10, pady=10)

    # Thông tin chi tiết
    frame_info = tk.Frame(frame_chi_tiet)
    frame_info.pack(fill=tk.X)

    # Cột 1
    frame_col1 = tk.Frame(frame_info)
    frame_col1.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)

    tk.Label(frame_col1, text="Mã hóa đơn:").grid(row=0, column=0, sticky=tk.W, pady=2)
    lbl_ma_hoa_don = tk.Label(frame_col1, text="", width=20)
    lbl_ma_hoa_don.grid(row=0, column=1, sticky=tk.W, pady=2)

    tk.Label(frame_col1, text="Ngày lập:").grid(row=1, column=0, sticky=tk.W, pady=2)
    lbl_ngay_lap = tk.Label(frame_col1, text="", width=20)
    lbl_ngay_lap.grid(row=1, column=1, sticky=tk.W, pady=2)

    tk.Label(frame_col1, text="Số tiền:").grid(row=2, column=0, sticky=tk.W, pady=2)
    lbl_so_tien = tk.Label(frame_col1, text="", width=20)
    lbl_so_tien.grid(row=2, column=1, sticky=tk.W, pady=2)

    # Cột 2
    frame_col2 = tk.Frame(frame_info)
    frame_col2.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)

    tk.Label(frame_col2, text="Mã Phòng:").grid(row=0, column=0, sticky=tk.W, pady=2)
    lbl_ma_p = tk.Label(frame_col2, text="", width=20)
    lbl_ma_p.grid(row=0, column=1, sticky=tk.W, pady=2)

    tk.Label(frame_col2, text="Người thuê:").grid(row=1, column=0, sticky=tk.W, pady=2)
    lbl_ho_ten = tk.Label(frame_col2, text="", width=20)
    lbl_ho_ten.grid(row=1, column=1, sticky=tk.W, pady=2)

    tk.Label(frame_col2, text="Giá phòng:").grid(row=2, column=0, sticky=tk.W, pady=2)
    lbl_gia_phong = tk.Label(frame_col2, text="", width=20)
    lbl_gia_phong.grid(row=2, column=1, sticky=tk.W, pady=2)

    # Cột 3
    frame_col3 = tk.Frame(frame_info)
    frame_col3.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)

    tk.Label(frame_col3, text="Mã hợp đồng:").grid(row=0, column=0, sticky=tk.W, pady=2)
    lbl_ma_hd = tk.Label(frame_col3, text="", width=20)
    lbl_ma_hd.grid(row=0, column=1, sticky=tk.W, pady=2)

    tk.Label(frame_col3, text="Thời gian BĐ:").grid(row=1, column=0, sticky=tk.W, pady=2)
    lbl_thoi_gian_bd = tk.Label(frame_col3, text="", width=20)
    lbl_thoi_gian_bd.grid(row=1, column=1, sticky=tk.W, pady=2)

    tk.Label(frame_col3, text="Thời gian KT:").grid(row=2, column=0, sticky=tk.W, pady=2)
    lbl_thoi_gian_kt = tk.Label(frame_col3, text="", width=20)
    lbl_thoi_gian_kt.grid(row=2, column=1, sticky=tk.W, pady=2)

    # Định nghĩa hàm lấy thông tin chi tiết khi click vào một hóa đơn
    def item_selected(event):
        for selected_item in tree_hoa_don.selection():
            item = tree_hoa_don.item(selected_item)
            record = item['values']
            lbl_ma_hoa_don.config(text=record[0])
            lbl_ma_hd.config(text=record[1])
            lbl_ma_p.config(text=record[2])
            lbl_ho_ten.config(text=record[4])
            lbl_so_tien.config(text=f"{record[5]} VNĐ")
            lbl_ngay_lap.config(text=record[6])
            lbl_gia_phong.config(text=f"{record[7]} VNĐ")

            # Lấy thông tin thời gian từ DB
            conn = ket_noi_csdl()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT ThoiGianBD, ThoiGianKT FROM HopDong WHERE MaHD = ?
            ''', (record[1],))
            hop_dong_info = cursor.fetchone()
            conn.close()

            if hop_dong_info:
                lbl_thoi_gian_bd.config(text=hop_dong_info[0])
                lbl_thoi_gian_kt.config(text=hop_dong_info[1])

    # Gắn sự kiện khi click vào một hàng trong bảng
    tree_hoa_don.bind('<<TreeviewSelect>>', item_selected)

    # Frame cho các nút chức năng mới
    frame_new_buttons = tk.Frame(parent_frame)
    frame_new_buttons.pack(side=tk.RIGHT, padx=10, pady=10)

    btn_issue_invoice = tk.Button(frame_new_buttons, text="Phát hành Hóa Đơn",
                                  command=lambda: issue_invoice(tree_hoa_don))
    btn_issue_invoice.pack(side=tk.LEFT, padx=5)

    btn_issue_multiple_invoices = tk.Button(frame_new_buttons, text="Phát hành Nhiều Hóa Đơn",
                                            command=lambda: issue_multiple_invoices(tree_hoa_don))
    btn_issue_multiple_invoices.pack(side=tk.LEFT, padx=5)

    return parent_frame


# Hàm lấy và hiển thị dữ liệu
def refresh_data(tree):
    for i in tree.get_children():
        tree.delete(i)

    conn = ket_noi_csdl()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT h.MaHoaDon, h.MaHD, p.MaPhong, n.CCCD, n.HoTen,
               h.SoTien, h.NgayLap, p.GiaPhong
        FROM HoaDon h
        JOIN HopDong hd ON h.MaHD = hd.MaHD
        JOIN PhongTro p ON hd.MaPhong = p.MaPhong
        JOIN NguoiThue n ON hd.CCCD = n.CCCD
        ORDER BY h.NgayLap DESC
    ''')

    records = cursor.fetchall()
    conn.close()

    for record in records:
        try:
            date_obj = datetime.strptime(record[6], '%Y-%m-%d')
            formatted_date = date_obj.strftime('%d/%m/%Y')
        except:
            formatted_date = record[6]

        tree.insert('', tk.END, values=(
            record[0], record[1], record[2], record[3], record[4],
            f"{record[5]:,.0f}", formatted_date, f"{record[7]:,.0f}"
        ))


# Hàm tìm kiếm hóa đơn
def search_bills(ma_hoa_don, tu_ngay, den_ngay, tree):
    for i in tree.get_children():
        tree.delete(i)

    conn = ket_noi_csdl()
    cursor = conn.cursor()

    query = '''
        SELECT h.MaHoaDon, h.MaHD, p.MaPhong, n.CCCD, n.HoTen,
               h.SoTien, h.NgayLap, p.GiaPhong
        FROM HoaDon h
        JOIN HopDong hd ON h.MaHD = hd.MaHD
        JOIN PhongTro p ON hd.MaPhong = p.MaPhong
        JOIN NguoiThue n ON hd.CCCD = n.CCCD
        WHERE 1=1
    '''
    params = []

    if ma_hoa_don:
        query += " AND h.MaHoaDon LIKE ?"
        params.append(f"%{ma_hoa_don}%")

    if tu_ngay is not None and den_ngay is not None:
        tu_ngay_str = tu_ngay.strftime('%Y-%m-%d')
        den_ngay_str = den_ngay.strftime('%Y-%m-%d')
        query += " AND h.NgayLap BETWEEN ? AND ?"
        params.extend([tu_ngay_str, den_ngay_str])

    query += " ORDER BY h.NgayLap DESC"

    cursor.execute(query, params)
    records = cursor.fetchall()
    conn.close()

    for record in records:
        try:
            date_obj = datetime.strptime(record[6], '%Y-%m-%d')
            formatted_date = date_obj.strftime('%d/%m/%Y')
        except:
            formatted_date = record[6]

        tree.insert('', tk.END, values=(
            record[0], record[1], record[2], record[3], record[4],
            f"{record[5]:,.0f}", formatted_date, f"{record[7]:,.0f}"
        ))
    

# Hàm thêm/sửa hóa đơn
def show_bill_form(edit_data, tree):
    bill_form = tk.Toplevel()
    bill_form.title("Thêm/Sửa Hóa Đơn")
    bill_form.geometry("400x600")

    # Tạo canvas và thanh cuộn
    canvas = tk.Canvas(bill_form)
    scrollbar = ttk.Scrollbar(bill_form, orient=tk.VERTICAL, command=canvas.yview)
    scrollable_frame = tk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    frame_form = scrollable_frame

    # Các trường nhập liệu
    tk.Label(frame_form, text="Mã hóa đơn:").grid(row=0, column=0, sticky=tk.W, pady=5, padx=5)
    entry_ma_hoa_don = tk.Entry(frame_form, width=30)
    entry_ma_hoa_don.grid(row=0, column=1, pady=5, padx=5)
    if not edit_data:
        entry_ma_hoa_don.insert(0, get_next_bill_id())
    entry_ma_hoa_don.config(state='readonly')

    tk.Label(frame_form, text="Ngày lập:").grid(row=1, column=0, sticky=tk.W, pady=5, padx=5)
    entry_ngay_lap = DateEntry(frame_form, width=28, background='darkblue',
                               foreground='white', date_pattern='dd/MM/yyyy')
    entry_ngay_lap.grid(row=1, column=1, pady=5, padx=5)

    tk.Label(frame_form, text="Mã hợp đồng:").grid(row=2, column=0, sticky=tk.W, pady=5, padx=5)
    hop_dong_combo = ttk.Combobox(frame_form, width=28)
    hop_dong_combo.grid(row=2, column=1, pady=5, padx=5)

    # Lấy danh sách hợp đồng từ CSDL
    conn = ket_noi_csdl()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT hd.MaHD, n.HoTen 
        FROM HopDong hd
        JOIN NguoiThue n ON hd.CCCD = n.CCCD
    ''')
    hop_dong_list = cursor.fetchall()
    conn.close()

    hop_dong_values = [f"{hd[0]} - {hd[1]}" for hd in hop_dong_list]
    hop_dong_ids = [hd[0] for hd in hop_dong_list]
    hop_dong_combo['values'] = hop_dong_values

    # Thông tin hợp đồng
    tk.Label(frame_form, text="Thông tin hợp đồng:").grid(row=3, column=0, columnspan=2, sticky=tk.W, pady=10, padx=5)
    lbl_info_nguoi_thue = tk.Label(frame_form, text="Người thuê: ")
    lbl_info_nguoi_thue.grid(row=4, column=0, columnspan=2, sticky=tk.W, pady=2, padx=5)
    lbl_info_phong = tk.Label(frame_form, text="Phòng: ")
    lbl_info_phong.grid(row=5, column=0, columnspan=2, sticky=tk.W, pady=2, padx=5)
    lbl_info_gia_phong = tk.Label(frame_form, text="Giá phòng: ")
    lbl_info_gia_phong.grid(row=6, column=0, columnspan=2, sticky=tk.W, pady=2, padx=5)

    # Chỉ số điện
    tk.Label(frame_form, text="Chỉ số điện đầu:").grid(row=7, column=0, sticky=tk.W, pady=5, padx=5)
    entry_dien_dau = tk.Entry(frame_form, width=15)
    entry_dien_dau.grid(row=7, column=1, pady=5, padx=5, sticky=tk.W)

    tk.Label(frame_form, text="Chỉ số điện cuối:").grid(row=8, column=0, sticky=tk.W, pady=5, padx=5)
    entry_dien_cuoi = tk.Entry(frame_form, width=15)
    entry_dien_cuoi.grid(row=8, column=1, pady=5, padx=5, sticky=tk.W)

    # Chỉ số nước
    tk.Label(frame_form, text="Chỉ số nước đầu:").grid(row=9, column=0, sticky=tk.W, pady=5, padx=5)
    entry_nuoc_dau = tk.Entry(frame_form, width=15)
    entry_nuoc_dau.grid(row=9, column=1, pady=5, padx=5, sticky=tk.W)

    tk.Label(frame_form, text="Chỉ số nước cuối:").grid(row=10, column=0, sticky=tk.W, pady=5, padx=5)
    entry_nuoc_cuoi = tk.Entry(frame_form, width=15)
    entry_nuoc_cuoi.grid(row=10, column=1, pady=5, padx=5, sticky=tk.W)

    # Tổng tiền
    tk.Label(frame_form, text="Tổng tiền (VNĐ):").grid(row=11, column=0, sticky=tk.W, pady=5, padx=5)
    entry_so_tien = tk.Entry(frame_form, width=30, state='readonly')
    entry_so_tien.grid(row=11, column=1, pady=5, padx=5)

    # Hàm tính tổng tiền
    def calculate_total():
        try:
            dien_dau = float(entry_dien_dau.get() or 0)
            dien_cuoi = float(entry_dien_cuoi.get() or 0)
            nuoc_dau = float(entry_nuoc_dau.get() or 0)
            nuoc_cuoi = float(entry_nuoc_cuoi.get() or 0)

            service_prices = get_service_prices()
            gia_phong_str = lbl_info_gia_phong.cget("text").split(":")[1].strip().replace(" VNĐ", "").replace(",", "")
            gia_phong = float(gia_phong_str) if gia_phong_str else 0

            dien_tieu_thu = dien_cuoi - dien_dau
            nuoc_tieu_thu = nuoc_cuoi - nuoc_dau

            tong_tien = (gia_phong +
                         dien_tieu_thu * service_prices['dien'] +
                         nuoc_tieu_thu * service_prices['nuoc'])

            entry_so_tien.config(state='normal')
            entry_so_tien.delete(0, tk.END)
            entry_so_tien.insert(0, f"{tong_tien:,.0f}")
            entry_so_tien.config(state='readonly')
        except ValueError:
            entry_so_tien.config(state='normal')
            entry_so_tien.delete(0, tk.END)
            entry_so_tien.insert(0, "0")
            entry_so_tien.config(state='readonly')

    # Hàm hiển thị thông tin hợp đồng khi chọn
    def on_hop_dong_selected(event):
        selected = hop_dong_combo.current()
        if selected >= 0:
            ma_hd = hop_dong_ids[selected]

            conn = ket_noi_csdl()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT p.GiaPhong, n.HoTen, p.MaPhong
                FROM HopDong hd
                JOIN NguoiThue n ON hd.CCCD = n.CCCD
                JOIN PhongTro p ON hd.MaPhong = p.MaPhong
                WHERE hd.MaHD = ?
            ''', (ma_hd,))
            info = cursor.fetchone()
            conn.close()

            if info:
                lbl_info_nguoi_thue.config(text=f"Người thuê: {info[1]}")
                lbl_info_phong.config(text=f"Phòng: {info[2]}")
                lbl_info_gia_phong.config(text=f"Giá phòng: {info[0]:,.0f} VNĐ")
                calculate_total()

    hop_dong_combo.bind("<<ComboboxSelected>>", on_hop_dong_selected)

    # Gắn sự kiện KeyRelease để tự động tính tổng tiền khi nhập chỉ số
    for entry in [entry_dien_dau, entry_dien_cuoi, entry_nuoc_dau, entry_nuoc_cuoi]:
        entry.bind("<KeyRelease>", lambda e: calculate_total())

    if edit_data:
        entry_ma_hoa_don.delete(0, tk.END)
        entry_ma_hoa_don.insert(0, edit_data[0])

        try:
            date_obj = datetime.strptime(edit_data[6], '%d/%m/%Y')
            entry_ngay_lap.set_date(date_obj)
        except:
            pass

        entry_so_tien.config(state='normal')
        entry_so_tien.insert(0, edit_data[5])
        entry_so_tien.config(state='readonly')

        for i, ma_hd in enumerate(hop_dong_ids):
            if ma_hd == edit_data[1]:
                hop_dong_combo.current(i)
                on_hop_dong_selected(None)
                break

    # Nút tính tổng tiền
    btn_calculate = tk.Button(frame_form, text="Tính tổng tiền", command=calculate_total)
    btn_calculate.grid(row=12, column=1, pady=5, sticky=tk.W)

    # Frame cho các nút
    frame_buttons = tk.Frame(frame_form)
    frame_buttons.grid(row=13, column=0, columnspan=2, pady=15)

    # Hàm lưu hóa đơn
    def save_bill():
        ma_hoa_don = entry_ma_hoa_don.get()
        ngay_lap = entry_ngay_lap.get_date().strftime('%Y-%m-%d')
        so_tien = float(entry_so_tien.get().replace(',', ''))
        selected = hop_dong_combo.current()

        if selected < 0:
            messagebox.showerror("Lỗi", "Vui lòng chọn hợp đồng!")
            return

        ma_hd = hop_dong_ids[selected]

        if not all([ma_hoa_don, ngay_lap, entry_dien_dau.get(), entry_dien_cuoi.get(),
                    entry_nuoc_dau.get(), entry_nuoc_cuoi.get()]):
            messagebox.showerror("Lỗi", "Vui lòng điền đầy đủ thông tin!")
            return

        conn = ket_noi_csdl()
        cursor = conn.cursor()

        try:
            if edit_data:
                cursor.execute('''
                    UPDATE HoaDon
                    SET NgayLap = ?, SoTien = ?, MaHD = ?
                    WHERE MaHoaDon = ?
                ''', (ngay_lap, so_tien, ma_hd, ma_hoa_don))
            else:
                cursor.execute('''
                    INSERT INTO HoaDon (MaHoaDon, NgayLap, SoTien, MaHD)
                    VALUES (?, ?, ?, ?)
                ''', (ma_hoa_don, ngay_lap, so_tien, ma_hd))

            conn.commit()
            messagebox.showinfo("Thông báo", "Lưu hóa đơn thành công!")
            bill_form.destroy()
            refresh_data(tree)
        except sqlite3.Error as e:
            messagebox.showerror("Lỗi Database", str(e))
        finally:
            conn.close()

    btn_save = tk.Button(frame_buttons, text="Lưu", width=10, command=save_bill)
    btn_save.pack(side=tk.LEFT, padx=5)

    btn_cancel = tk.Button(frame_buttons, text="Hủy", width=10, command=bill_form.destroy)
    btn_cancel.pack(side=tk.LEFT, padx=5)

def edit_selected_bill(tree):
    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("Cảnh báo", "Vui lòng chọn hóa đơn cần sửa!")
        return

    selected_item = selected_items[0]
    values = tree.item(selected_item, 'values')
    show_bill_form(values, tree)

def delete_selected_bill(tree):
    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("Cảnh báo", "Vui lòng chọn hóa đơn cần xóa!")
        return

    result = messagebox.askyesno("Xác nhận", "Bạn có chắc chắn muốn xóa hóa đơn này?")
    if not result:
        return

    selected_item = selected_items[0]
    ma_hoa_don = tree.item(selected_item, 'values')[0]

    conn = ket_noi_csdl()
    cursor = conn.cursor()

    try:
        cursor.execute('DELETE FROM HoaDon WHERE MaHoaDon = ?', (ma_hoa_don,))
        conn.commit()
        messagebox.showinfo("Thông báo", "Xóa hóa đơn thành công!")
        refresh_data(tree)
    except sqlite3.Error as e:
        messagebox.showerror("Lỗi Database", str(e))
    finally:
        conn.close()


# Hàm xuất dữ liệu ra Excel
def export_to_excel(tree):
    data = []
    columns = ['Mã Hóa Đơn', 'Mã Hợp Đồng', 'Mã Phòng', 'CCCD', 'Họ Tên Người Thuê', 'Số Tiền', 'Ngày Lập', 'Giá Phòng']

    for item in tree.get_children():
        values = tree.item(item, 'values')
        data.append(list(values))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hóa Đơn"

    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)

    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    try:
        file_path = tk.filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel Files', '*.xlsx')],
            title='Lưu file Excel'
        )

        if file_path:
            wb.save(file_path)
            messagebox.showinfo("Thông báo", f"Đã xuất dữ liệu ra file: {file_path}")
    except Exception as e:
        messagebox.showerror("Lỗi", f"Có lỗi khi xuất Excel: {str(e)}")

def issue_invoice(tree):
    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("Cảnh báo", "Vui lòng chọn hóa đơn cần phát hành!")
        return

    data = tree.item(selected_items[0], 'values')
    try:
        conn = ket_noi_csdl()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT h.MaHoaDon, h.MaHD, n.CCCD, n.HoTen,
                   h.SoTien, h.NgayLap, p.GiaPhong
            FROM HoaDon h
            JOIN HopDong hd ON h.MaHD = hd.MaHD
            JOIN NguoiThue n ON hd.CCCD = n.CCCD
            JOIN PhongTro p ON hd.MaPhong = p.MaPhong
            WHERE h.MaHoaDon = ?
            ORDER BY h.NgayLap DESC
        ''', (data[0],))

        records = cursor.fetchall()
        conn.close()

        if not records:
            messagebox.showinfo("Thông báo", "Không có dữ liệu để xuất.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active

        ws['B2'] = 'Công Ty TNHH ABC'
        ws['B3'] = 'Địa chỉ: 123 Đường XYZ, Quận 1, TP.HCM'
        ws['B4'] = 'Điện thoại: 01234567890'

        ws['B6'] = 'Mã hóa đơn'
        ws['C6'] = data[0]
        ws['B7'] = 'Mã hợp đồng'
        ws['C7'] = data[1]
        ws['B8'] = 'Họ tên'
        ws['C8'] = data[4]
        ws['B9'] = 'Ngày lập'
        ws['C9'] = data[6]
        ws['B10'] = 'Giá phòng'
        ws['C10'] = data[7]
        ws['B11'] = 'Số tiền'
        ws['C11'] = data[5]
        ws['B13'] = 'Tổng tiền:'
        ws['C13'] = f"{data[5]} đ"

        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15

        for row in range(1, 13):
            for col in ['A', 'B', 'C']:
                if ws[f'{col}{row}'].value:
                    ws[f'{col}{row}'].font = Font(name='Arial', size=11)
                    ws[f'{col}{row}'].alignment = Alignment(vertical='center')

        ws['A1'].font = Font(bold=True, size=12)

        for row in range(1, 15):
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row}']
                border = Border(
                    left=Side(style=None),
                    right=Side(style=None),
                    top=Side(style=None),
                    bottom=Side(style=None)
                )
                if row == 1:
                    border.top = Side(style='thick')
                if row == 14:
                    border.bottom = Side(style='thick')
                if col == 'A':
                    border.left = Side(style='thick')
                if col == 'D':
                    border.right = Side(style='thick')
                cell.border = border

        wb.save('hoa_don.xlsx')
        messagebox.showinfo("Thông báo", "Xuất hóa đơn thành công! File hoa_don.xlsx đã được tạo.")

    except Exception as e:
        messagebox.showerror("Lỗi", f"Đã xảy ra lỗi: {e}")

def issue_multiple_invoices(tree):
    messagebox.showinfo("Thông báo", "Chức năng phát hành nhiều hóa đơn chưa được triển khai!")


def run_test():
    root = tk.Tk()
    root.title("Quản lý hóa đơn")
    root.geometry("1000x600")
    quan_ly_hoa_don(root)
    root.mainloop()

if __name__ == "__main__":
    run_test()